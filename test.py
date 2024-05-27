from platform import libc_ver
from xml.dom.minidom import Identified
import openpyxl
import time
import re
import os
import sys 
from openpyxl.utils.cell import get_column_letter
from datetime import datetime, date
import pathlib
z = []
cle = date(2022, 10, 18)   

def funct():
     for fmt in ('%d-%m-%Y', '%d.%m.%Y', '%d/%m/%Y', '%m/%d/%y'):
        try:
            fname = datetime.strptime(mat_str.group(), fmt).date()
            print(fname)
            
        except:
            pass

t = funct()

   
path = input("Enter the path:")
n_path = input("enter new path")



os.chdir(path)

excel_sheet = os.listdir('.')

for list in range(0, len(excel_sheet)):
    wb = openpyxl.load_workbook(excel_sheet[list])
    book = wb.active
    x =book.max_column
    y = x - 2
    grt = get_column_letter(y)
    c = str(grt) + "1"
    gt_string = book[c]
    tst2 = str(excel_sheet[list])
    print(excel_sheet[list])
    tst = str(gt_string.value)
    tst_str = tst

    if re.search(r'\d{2}.\d{1}.\d{4}', tst_str) != None:
        mat_str =re.search(r'\d{2}.\d{1}.\d{4}', tst_str)
        time.sleep(0.5)
        for fmt in ('%d-%m-%Y', '%d.%m.%Y', '%d/%m/%Y'):
            try:
                fname = datetime.strptime(mat_str.group(), fmt).date()
                print(fname)
            except:
                pass
        

    elif re.search(r'\d{2}.\d{2}.\d{4}', tst_str) != None:
        mat_str = re.search(r'\d{2}.\d{2}.\d{4}', tst_str)
        time.sleep(0.5)
        for fmt in ('%d-%m-%Y', '%d.%m.%Y', '%d/%m/%Y'):
            try:
                fname = datetime.strptime(mat_str.group(), fmt).date()
                print(fname)
            except:
                pass
        
        
    elif re.search(r'\d{2}.\d{1}.\d{2}', tst_str) != None:
        mat_str = re.search(r'\d{2}.\d{1}.\d{2}', tst_str)
        time.sleep(0.5)
        for fmt in ('%d-%m-%Y', '%d.%m.%Y', '%d/%m/%Y',):
            try:
                fname = datetime.strptime(mat_str.group(), fmt).date()
                print(fname)
            except:
                pass
        
    elif re.search(r'\d{2}.\d{2}.\d{2}', tst_str) != None:
        mat_str = re.search(r'\d{2}.\d{2}.\d{2}', tst_str)
        time.sleep(0.5)
        #res = datetime.strptime(mat_str.group()).date()
        for fmt in ('%d-%m-%Y', '%d.%m.%Y', '%d/%m/%Y'):
            try:
                fname = datetime.strptime(mat_str.group(), fmt).date()
                print(fname)
            except:
                pass

    elif re.search(r'\d{1}.\d{2}.\d{4}', tst_str) != None:
        mat_str = re.search(r'\d{1}.\d{2}.\d{4}', tst_str)
        time.sleep(0.5)
        #res = datetime.strptime(mat_str.group()).date()
        for fmt in ('%d-%m-%Y', '%d.%m.%Y', '%d/%m/%Y'):
            try:
                fname = datetime.strptime(mat_str.group(), fmt).date()
                print(fname)
            except:
                pass
    
    else:
        print(book['x1'].value)




    

    





    
    
        


    



