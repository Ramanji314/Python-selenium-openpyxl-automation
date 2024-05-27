import openpyxl
import re
import os
import sys
import xlrd
import time
import shutil
 

path = input("provide a path")

filler = []

os.chdir(path)

main = os.listdir('.')

print(len(main))

count = 0
for i in range(0, len(main)):
    title = str(main[i])
    title1 = title[0:21]
    name2 = re.split('-', title1)
    for name3 in name2:
        name3.replace(" ", "")
        if len(name3) == 7 or len(name3) == 9:
            if name3 not in filler:
                filler.append(name3)
                name4 = name3
            else:
                name4 = name3+ "-1"
                filler.append(name4)
            wb = openpyxl.load_workbook(main[i])
            print(wb.get_sheet_names())
            wn = wb.get_sheet_by_name("Quiz information")
            wb.remove_sheet(wn)
            wb.get_sheet_by_name("Quiz structure analysis")
            sheet = wb.active
            print(sheet.max_column)
            print(sheet.max_row)
            for mi in range(1, sheet.max_column+1):
                for ls in range(1, sheet.max_row+1):
                    s = sheet.cell(row = ls, column = mi).value
                    if s == None:
                        sheet.cell(row = ls, column = mi).value = "-"
                        count += 1
            print(count)

            wb.save("C:\\Users\\raman\\statistics\\" +name4  +".xlsx")
    time.sleep(0.5)
print(filler)







