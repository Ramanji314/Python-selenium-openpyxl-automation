import re
import openpyxl
import os
import sys
from openpyxl.utils.cell import get_column_letter
import time
import bipdict
import selenium
from time import time
from cmath import exp
from time import time
from tkinter.messagebox import NO
from turtle import delay
from xml.dom import DOMSTRING_SIZE_ERR
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait 
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import os
import sys
import shutil
import re
from selenium.webdriver.support import expected_conditions as EC
import time
from bipdict import subjects
from selenium.webdriver.support.select import Select



path = input("provide the path")

os.chdir(path)

name = os.listdir('.')



bit_dict ={}

final_name = ""
final_file = []

downloaded = []
not_downloaded = []


for i in range(0, len(name)):
    name2 = str(name[i])
    name3 = re.split('-| ', name2)
    for nam in name3:
        if len(nam) == 7 or len(nam) == 9:
           final_name = nam
           if final_name not in final_file:
               final_name = final_name
               final_file.append(final_name)
           else:
               final_name = final_name +"1"
               final_file.append(final_name)
    wb = openpyxl.load_workbook(name[i])
    sheet = wb.active
    column = sheet.max_column
    col = column-2
    grt = get_column_letter(col)
    grt2 = str(grt)
    bt = grt2 +"1"
    get_name = sheet[bt]
    get_name2 = get_name.value
    le = len(get_name2)
    str2 = get_name2[6:le-7]
    bit_dict[final_name] = str2
  
    
       

print(bit_dict)

J = 1
driver = webdriver.Chrome(executable_path="D:\\python files for moodle work\\chromedriver.exe")
driver.get('https://accounts.google.com/ServiceLogin/identifier?continue=https%3A%2F%2Fmail.google.com&sacu=1&passive=1209600&hl=en&acui=0&flowName=GlifWebSignIn&flowEntry=ServiceLogin&cid=1&TL=AKqFyY9fS1dRw8Qp3lKdBHb1jqC_m2CKQe6WwEU_FRewwf9xyEqNOvyklJyp7Jt8')
driver.implicitly_wait(100)

loginBox = driver.find_element_by_xpath('//*[@id ="identifierId"]')
loginBox.send_keys("email")

nextButton = driver.find_element_by_xpath('//*[@id ="identifierNext"]')
nextButton.click()

passWordBox = driver.find_element_by_xpath('//*[@id ="password"]/div[1]/div / div[1]/input')
passWordBox.send_keys("password")


nextButton = driver.find_element_by_xpath('//*[@id ="passwordNext"]')
nextButton.click()
time.sleep(10)

driver.implicitly_wait(5000)

driver.execute_script("window.open('');")
driver.switch_to.window(driver.window_handles[J])

driver.get('https://moodle.bitsathy.ac.in/grade/export/xls/index.php?id=10793')
self_in = driver.find_element_by_xpath("/html[1]/body[1]/div[4]/div[1]/div[1]/div[1]/section[1]/div[2]/div[2]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[3]/div[1]/a[1]")
self_in.click()

for x in bit_dict:
    try:
        c_name = x
        l_name = bit_dict[x]
        if len(c_name) == 8 or len(c_name) == 10:
            length = len(c_name)
            c_name2 = c_name[0:length-1]
            print(c_name2)
        else:
            c_name2 = c_name 
            print(c_name2)
        num3 = subjects.get(c_name2)
        path = "https://moodle.bitsathy.ac.in/course/view.php?id="+ num3
        driver.get(path)
        time.sleep(3)
        yyy = driver.find_element_by_partial_link_text(l_name)
        time.sleep(1)
        yyy.click()
        lr = driver.current_url
        les = re.split('=', lr)
        gt =les[1]
        gt2 = str(gt)
        path2 = "https://moodle.bitsathy.ac.in/mod/quiz/report.php?id=" +gt2 +"&mode=statistics"
        driver.get(path2)
        driver.implicitly_wait(15)
        ddl = driver.find_element_by_xpath("(//select[@class='form-control custom-select mr-1'])[2]")
        driver.implicitly_wait(15)
        ee = Select(ddl)
        ee.select_by_value("excel")
        time.sleep(2)
        driver.implicitly_wait(15)
        try:
            xyz = driver.find_element_by_xpath("(//button[@class='btn btn-secondary'])[2]")
            xyz.click()
            driver.implicitly_wait(100)
        except:
            xyz = driver.find_element_by_xpath("(//button[@class='btn btn-secondary'])[3]")
            xyz.click()
        print("Sucessfully downloaded", x)
        driver.implicitly_wait(15)
        downloaded.append(x)
    except:
        print("the file should be downloaded for:", x)
        not_downloaded.append(x)



print("the total number of FA Downloaded:", len(downloaded))
print("the total number of FA not Downloaded:", len(not_downloaded))



    











    


    


    




        


