from operator import le
from time import time
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import os 
import sys
import re
import time
from bipdict import souce
from selenium.webdriver.support.select import Select


x =1 
path = input("provide a path:")

os.chdir(path)

excelsheets = os.listdir('.')

min_len = 7
let = "FA"

y = []
z = []
ident = []

driver = webdriver.Chrome(executable_path="D:\\python files for moodle work\\chromedriver.exe")
driver.get('https://accounts.google.com/ServiceLogin/identifier?continue=https%3A%2F%2Fmail.google.com&sacu=1&passive=1209600&hl=en&acui=0&flowName=GlifWebSignIn&flowEntry=ServiceLogin&cid=1&TL=AKqFyY9fS1dRw8Qp3lKdBHb1jqC_m2CKQe6WwEU_FRewwf9xyEqNOvyklJyp7Jt8')
driver.implicitly_wait(100)

loginBox = driver.find_element_by_xpath('//*[@id ="identifierId"]')
loginBox.send_keys("edamakantiramanjireddy@bitsathy.ac.in")

nextButton = driver.find_element_by_xpath('//*[@id ="identifierNext"]')
nextButton.click()

passWordBox = driver.find_element_by_xpath('//*[@id ="password"]/div[1]/div / div[1]/input')
passWordBox.send_keys("Ramanji.1234")

nextButton = driver.find_element_by_xpath('//*[@id ="passwordNext"]')
nextButton.click()
time.sleep(10)

driver.execute_script("window.open('');")
driver.switch_to.window(driver.window_handles[1])

driver.get('https://bip.bitsathy.ac.in/nova/resources/courses/812')

time.sleep(3)

clr_button = driver.find_element_by_xpath("//div[@class='w1I7fb']")
clr_button.click()

time.sleep(3)

in_button = driver.find_element_by_xpath("//a[@class='w-full h- btn btn-default btn-primary hover:bg-primary-dark text-center']")
in_button.click()

time.sleep(3)

for i in range(0, len(excelsheets)):
    wb = openpyxl.load_workbook(excelsheets[i])
    sheet = wb.active
    n = sheet.max_column
    sheet.delete_cols(idx=n-1, amount=2)
    sheet.delete_cols(idx=1, amount=2)
    wb.save(excelsheets[i]) 
    col_length = sheet.max_column
    if col_length == 2:
        sheet['A1'] = "Email address"
        sheet['B1'] = "Grade"
        wb.save(excelsheets[i])
    elif col_length == 3:
        sheet.delete_cols(idx=2, amount=col_length-2)
        sheet['A1'] = "Email address"
        sheet['B1'] = "Grade"
        wb.save(excelsheets[i])
    elif col_length > 3:
        sheet.delete_cols(idx=2, amount=col_length-2)
        sheet['A1'] = "Email address"
        sheet['B1'] = "Grade"
        wb.save(excelsheets[i])
    name = str(excelsheets[i])
    gog = re.split(' |-|_', name)   
    min_word = ""
    fa_name = col_length - 1
    f_n = let+ str(fa_name)
    for ele in gog:
        if len(ele) == min_len or len(ele)==9:
            min_word = ele
            if min_word not in ident:
                ident.append(min_word)
                name4 = min_word+ ".xlsx"
                
            else:
                name4 = min_word+ "-1"+ ".xlsx"
            name5 = str(name4)
            print(name5)
            code = souce.get(ele)
            if code == None:
                print("the subject code is not added to library" +str(code))
            else:
                try:
                    path = "https://bip.bitsathy.ac.in/nova/resources/courses/"+ code
                    path2 = "C:/Users/raman/Downloads/"+ name
                    path3 = "C:/Users/raman/Statistics/"+ name4
                    driver.get(path)
                    time.sleep(2)
                    btn1 = driver.find_element_by_xpath('/html[1]/body[1]/div[1]/div[1]/div[2]/div[2]/div[2]/div[5]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/a[1]')
                    btn1.click()
                    time.sleep(4)
                    f1_input = driver.find_element_by_xpath("//input[@id='name']")
                    f1_input.send_keys(f_n)
                    time.sleep(2)
                    eee = driver.find_element_by_xpath('//*[@id="fa_type"]')
                    sel = Select(eee)
                    sel.select_by_value("Regular")
                    time.sleep(1)
                    f2_input = driver.find_element_by_xpath("//input[@id='mark_max']")
                    f2_input.send_keys("10")
                    time.sleep(2)
                    f3_input = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/div[2]/form/div[1]/div/div[5]/div[2]/div/input[2]")
                    f3_input.click()
                    f3_output = driver.find_element_by_xpath("//span[@aria-label='August 12, 2023']")
                    f3_output.click()
                    fin_btn = driver.find_element_by_xpath('/html[1]/body[1]/div[1]/div[1]/div[2]/div[2]/div[2]/form[1]/div[2]/button[2]/span[1]')
                    fin_btn.click()
                    driver.implicitly_wait(20)
                    ddd = driver.find_element_by_xpath("//select[@class='form-control form-select mr-2']")
                    drop = Select(ddd)
                    drop.select_by_value("import-fa-grades")
                    time.sleep(1)
                    sec_btn = driver.find_element_by_xpath("//button[@title='Run Action']//*[name()='svg']")
                    sec_btn.click()
                    driver.find_element_by_xpath('//*[@id="file-fas-file"]').send_keys(path2)
                    time.sleep(3)
                    reddy2 = driver.find_element_by_xpath("//span[normalize-space()='Run Action']")
                    reddy2.click()
                    driver.implicitly_wait(30)
                    time.sleep(4)
                    ddd4 = driver.find_element_by_xpath("//select[@class='form-control form-select mr-2']")
                    drop1 = Select(ddd4)
                    drop1.select_by_value("import-moodle-quiz-structure_analysiss")
                    time.sleep(2)
                    sec_btn = driver.find_element_by_xpath("//button[@title='Run Action']//*[name()='svg']")
                    sec_btn.click()
                    driver.find_element_by_xpath('//*[@id="file-fas-file"]').send_keys(path3)
                    time.sleep(3)
                    reddy3 = driver.find_element_by_xpath("//span[normalize-space()='Run Action']")
                    reddy3.click()
                    time.sleep(1)
                    print(f_n)
                    print(min_word)
                    driver.implicitly_wait(5)
                    y.append(ele)
        
                except:
                    z.append(ele)
                    print("file should be uploaded manually" +str(ele))


driver.implicitly_wait(10)
driver.close()
print("the total number of files uploaded is:" +str(len(y)))
print("the total number of file should manually uplaod:" +str(len(z)))



    


    

    






    


